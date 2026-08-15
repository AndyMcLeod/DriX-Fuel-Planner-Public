"""DriX 8 Endurance sheet — EM2040 gondola, measured curves.

Matches the layout of the 'Endurance Data' tab in the Hourly Ops Log
(previously EM712 values): RPM | SOG-M/S | SOG-KNOTS | LPH | PCT/HR | NM/LITER,
nominal-values footnote, conversion cell, column chart of NM/L vs knots.
"""
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

import sys
# Default alongside the other documents in docs/; pass a path to write elsewhere.
OUT = Path(sys.argv[1]) if len(sys.argv) > 1 else (
    Path(__file__).resolve().parent.parent / 'docs' / 'DriX8_Endurance_EM2040.xlsx')

# Through-origin variants of the measured EM2040 laws, refitted from the same
# cruise bins as the planner's laws plus the measured loiter anchor, constrained
# so 0 rpm -> 0 L/h and 0 kt EXACTLY. This is the physics of the drivetrain:
# direct drive, shaft rpm = engine rpm (the trawling motor is no longer used),
# so zero shaft rpm means the engine is stopped.
#
# ⚠ THIS READS THE LIVE FIT, and it did not always. Until 2026-08-15 it loaded a
# dated snapshot (em2040_fit_2026-08-09.json), so the 16-session refit rebuilt
# every other document while this sheet came out BYTE-IDENTICAL — the ops-facing
# table silently pinned to superseded coefficients. It was caught by `git status`
# showing the .xlsx modified and the .csv beside it untouched, which is only
# possible if the numbers never moved. Read the live file, or a refit does not
# reach the sheet an operator actually carries.
import numpy as np
_LIVE = Path(__file__).with_name('rosbags') / 'em2040_fit.json'
if not _LIVE.exists():
    raise SystemExit(f'{_LIVE} missing — run tools/fit_em2040.py first')
FIT = json.load(open(_LIVE))
# Gauge scale read from the model file so this sheet's footnote cannot drift
# from the planner. (encoding='utf-8' on purpose — this machine defaults to
# cp1252 and silently mojibakes the em-dashes in model.json.)
with open(Path(__file__).resolve().parent.parent / 'model.json', encoding='utf-8') as _fh:
    _GC = json.load(_fh)['gauge_calibration']
GAUGE_LPP = _GC['l_per_point']
GAUGE_SIG = _GC['l_per_point_sigma']
GAUGE_BAND = _GC['band_pct']
# The reserve floor is policy and it moves; read it rather than writing it out.
with open(Path(__file__).resolve().parent.parent / 'model.json', encoding='utf-8') as _fh:
    _M = json.load(_fh)
RES_PCT = _M['reserve']['default_fraction'] * 100.0
_TANK = _M['tank_volume']['litres']
_R = np.array([b['rpm'] for b in FIT['bins']])
_K = np.array([b['kt'] for b in FIT['bins']])
_L = np.array([b['lph'] for b in FIT['bins']])
_W = np.array([min(b['n'], 3600) for b in FIT['bins']], float)
# The loiter anchor is read, not written out: it moved 0.95 -> 1.05 at the
# Aug-15 refit and a hardcoded copy here would have pulled the whole
# through-origin curve toward a burn the vehicle no longer has.
_LOITER = _M['gondolas']['options']['em2040']['loiter']
_Rf = np.append(_R, _LOITER['rpm'])              # loiter anchor (measured)
_Lf = np.append(_L, _LOITER['lph'])
_Wf = np.append(_W, 3600.0)
_A = np.vstack([_Rf, _Rf ** 2, _Rf ** 3]).T
_Wm = np.diag(_Wf)
A1, A2, A3 = np.linalg.solve(_A.T @ _Wm @ _A, _A.T @ _Wm @ _Lf)
SM0 = float(np.sum(_W * _R * _K) / np.sum(_W * _R * _R))   # kt = SM0 * rpm

# How far these through-origin variants sit from the planner's own laws, MEASURED
# rather than asserted. It used to be a sentence claiming "<= 3.7% in-window,
# 0.3% at 8 kt" — true of one fit and quietly wrong after any other.
_G2040 = _M['gondolas']['options']['em2040']
_PS, _PF = _G2040['speed_vs_rpm'], _G2040['fuel_vs_rpm']
_win = np.arange(_PF['valid_rpm_min'], _PF['valid_rpm_max'] + 1, 10.0)
_prim = _PF['q0'] + _PF['q1'] * _win + _PF['q2'] * _win ** 2
_thru = A1 * _win + A2 * _win ** 2 + A3 * _win ** 3
_dev = np.abs(_thru / _prim - 1.0) * 100
_AGREE_MAX = float(np.max(_dev))
_AGREE_AT = float(_win[int(np.argmax(_dev))])
# Where it settles: the through-origin constraint pulls hardest at the bottom of
# the window, toward the loiter anchor, and is negligible over the rest.
_above = _win > 1600
_AGREE_REST = float(np.max(_dev[_above])) if _above.any() else _AGREE_MAX
_r8 = (8.0 - _PS['b']) / _PS['m']
_AGREE_8KT = abs((A1 * _r8 + A2 * _r8 ** 2 + A3 * _r8 ** 3)
                 / (_PF['q0'] + _PF['q1'] * _r8 + _PF['q2'] * _r8 ** 2) - 1.0) * 100


def lph_at(rpm):
    return A1 * rpm + A2 * rpm * rpm + A3 * rpm ** 3


def kt_at(rpm):
    return SM0 * rpm


V_LO, V_HI = 1400, 3100
MS_KT = 1.9438444924406046

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Endurance Data EM2040'

BOLD = Font(name='Arial', size=10, bold=True)
BODY = Font(name='Arial', size=10)
NOTE = Font(name='Arial', size=9, bold=True)
SMALL = Font(name='Arial', size=9, italic=True)

for col, w in zip('ABCDEF', (10, 14, 14, 12, 13, 12)):
    ws.column_dimensions[col].width = w

ws['A1'] = 'DriX 8 Endurance with EM2040 Gondola'
ws['A1'].font = BOLD

headers = ['RPM', 'SOG - M/S*', 'SOG - KNOTS*', 'LPH', 'PCT / HR', 'NM / LITER']
for i, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=i, value=h)
    c.font = BOLD

RPMS = [0, 1000, 1250, 1500, 1750, 2000, 2250, 2500, 2750, 3000]
R0 = 4
for i, rpm in enumerate(RPMS):
    r = R0 + i
    ws[f'A{r}'] = rpm
    ws[f'A{r}'].font = BODY
    kt = kt_at(rpm)
    lph = lph_at(rpm)
    flag = '' if (V_LO <= rpm <= V_HI or rpm == 0) else '†'
    ws[f'B{r}'] = round(kt / MS_KT, 9)                       # SOG m/s (source col)
    ws[f'C{r}'] = f'=B{r}*$B$15'                             # knots, as the template
    ws[f'D{r}'] = round(lph, 9)
    ws[f'E{r}'] = f'=D{r}*100/$B$16'                          # % of tank per hour
    ws[f'F{r}'] = f'=IF(D{r}>0,C{r}/D{r},0)'
    for col in 'BCDEF':
        ws[f'{col}{r}'].font = BODY
        ws[f'{col}{r}'].number_format = '0.000'
    if flag:
        g = ws.cell(row=r, column=7, value=flag)
        g.font = SMALL

R1 = R0 + len(RPMS) - 1                                       # row 13

ws['A14'] = '* NOTE Actual SOG will vary with conditions - these are nominal values ONLY'
ws['A14'].font = NOTE
ws['A15'] = 'M/S / Knot'
ws['A15'].font = BOLD
ws['B15'] = round(MS_KT, 5)
ws['B15'].font = BODY
ws['A16'] = 'Tank L'
ws['A16'].font = BOLD
ws['B16'] = 250.0
ws['B16'].font = BODY

prov = [
    f'Source: DriX-8 MCAP logs (EM2040 fitted) — {FIT["cruise_hours"]:.2f} h '
    f'steady cruise over {_R.min():.0f}-{_R.max():.0f} rpm, flow meter vs thruster RPM.',
    f'Fuel law (through origin) L/H = {A1:.4e}·RPM {A2:+.4e}·RPM² {A3:+.4e}·RPM³,',
    f'anchored by (0,0) and the measured {_LOITER["lph"]:.2f} L/H station-keeping figure '
    f'at ~{_LOITER["rpm"]:.0f} rpm.',
    f'Speed law (through origin) KNOTS = {SM0:.6f}·RPM  (SOG-based, ±5-8% tidal).',
    f'Against the planner\'s own laws (model.json): {_AGREE_8KT:.1f}% at 8 kt, '
    f'≤{_AGREE_REST:.1f}% above 1600 rpm, worst {_AGREE_MAX:.1f}% at '
    f'{_AGREE_AT:.0f} rpm where the through-origin constraint bends toward idle.',
    'Drivetrain: direct drive, shaft rpm = engine rpm (trawling motor unused). 0 rpm = engine OFF = 0 L/H.',
    'The engine idles at ~1000 rpm (≈0.95 L/H — the 1000 row) and engages at ~1100 minimum:',
    'the 0-1000 rpm band is not an operating region; the curve merely passes through it.',
    f'† outside the cruise-measured {V_LO}-{V_HI} rpm window (extrapolated toward the idle/origin anchors).',
    'PCT/HR is computed on the 250 L nominal tank (cell B16), which implies 2.50 L per indicated point.',
    f'THE GAUGE IS MEASURED AT {GAUGE_LPP:.2f} ± {GAUGE_SIG:.2f} L/POINT ({GAUGE_BAND[0]:.0f}-{GAUGE_BAND[1]:.0f}% band) — '
    f'{(2.50 - GAUGE_LPP) / GAUGE_SIG:.0f}σ below that — so the needle',
    f'falls about {2.50 / GAUGE_LPP - 1:.0%} FASTER than this column shows. Set B16 = {100 * GAUGE_LPP:.0f} to read PCT/HR on the measured scale.',
    f'The {RES_PCT:.0f}% floor is a needle position: on the measured scale a mission may spend about '
    f'{(100 - RES_PCT) * GAUGE_LPP:.0f} L before reaching it, against {_TANK * (1 - RES_PCT / 100):.0f} L on a linear tank.',
    'Gauge non-linearity is UNRESOLVED (per-day spread only 1.5σ), not established; every reading is from the top third',
    'of the tank and the reserve band itself has never been calibrated. See docs/DriX8_Fuel_Gauge_Linearity.docx.',
    'Full derivation: DriX_Fuel_Efficiency_Report §5.5 and the AndyMcLeod/DriX-Fuel-Planner repo.',
]
for i, line in enumerate(prov):
    ws.cell(row=17 + i, column=1, value=line).font = SMALL

# ---- whole-knot chart source: bars land on integer knots for readability ----
ws['H3'] = 'KNOTS'
ws['I3'] = 'RPM*'
ws['J3'] = 'LPH'
ws['K3'] = 'NM / LITER'
for c in ('H3', 'I3', 'J3', 'K3'):
    ws[c].font = BOLD
for col, w in (('H', 9), ('I', 9), ('J', 10), ('K', 12)):
    ws.column_dimensions[col].width = w
KTS = list(range(4, 12))
K0 = 4
for i, v in enumerate(KTS):
    r = K0 + i
    rpm_v = v / SM0
    lph_v = lph_at(rpm_v)
    ws[f'H{r}'] = v
    ws[f'H{r}'].number_format = '0'
    ws[f'I{r}'] = round(rpm_v, 0)
    ws[f'I{r}'].number_format = '0'
    ws[f'J{r}'] = round(lph_v, 9)
    ws[f'J{r}'].number_format = '0.000'
    ws[f'K{r}'] = f'=H{r}/J{r}'
    ws[f'K{r}'].number_format = '0.000'
    for col in 'HIJK':
        ws[f'{col}{r}'].font = BODY
    if not (V_LO <= rpm_v <= V_HI):
        ws.cell(row=r, column=12, value='†').font = SMALL
K1 = K0 + len(KTS) - 1
ws[f'H{K1+1}'] = ('Chart source — whole-knot steps from the same measured laws; '
                  '† below the 1400 rpm window.')
ws[f'H{K1+1}'].font = SMALL

ch = BarChart()
ch.type = 'col'
ch.title = 'NM / LITER vs. SOG - KNOTS'
ch.y_axis.title = 'NM / LITER'
ch.x_axis.title = 'SOG - KNOTS'
ch.height, ch.width = 9.5, 15
ch.legend = None
ch.add_data(Reference(ws, min_col=11, min_row=K0, max_row=K1), titles_from_data=False)
ch.set_categories(Reference(ws, min_col=8, min_row=K0, max_row=K1))
ch.y_axis.delete = False
ch.x_axis.delete = False
ch.gapWidth = 40
from openpyxl.chart.shapes import GraphicalProperties
ch.series[0].graphicalProperties = GraphicalProperties(solidFill='4285F4')
ch.series[0].varyColors = False
ws.add_chart(ch, 'B18')

wb.save(OUT)
print('written:', OUT)

# ---- CSV twin: values-only path for Google Sheets imports that mangle xlsx ----
import csv
csv_out = OUT.with_suffix('.csv')
rows = [['DriX 8 Endurance with EM2040 Gondola'], [],
        ['RPM', 'SOG - M/S*', 'SOG - KNOTS*', 'LPH', 'PCT / HR', 'NM / LITER', '']]
for rpm in RPMS:
    kt_, l_ = kt_at(rpm), lph_at(rpm)
    flag = '' if (V_LO <= rpm <= V_HI or rpm == 0) else 'dagger'
    rows.append([rpm, round(kt_ / MS_KT, 6), round(kt_, 6), round(l_, 6),
                 round(l_ * 100 / 250, 6), round(kt_ / l_, 6) if l_ > 0 else 0, flag])
rows += [['* NOTE Actual SOG will vary with conditions - these are nominal values ONLY'],
         ['M/S / Knot', round(MS_KT, 5)], ['Tank L', 250], [],
         ['Whole-knot chart source'], ['KNOTS', 'RPM*', 'LPH', 'NM / LITER']]
for v in KTS:
    r_ = v / SM0
    l_ = lph_at(r_)
    rows.append([v, round(r_), round(l_, 6), round(v / l_, 6),
                 'dagger' if not V_LO <= r_ <= V_HI else ''])
rows += [[]] + [[line] for line in prov]
with open(csv_out, 'w', newline='', encoding='utf-8-sig') as fh:
    csv.writer(fh).writerows(rows)
print('written:', csv_out)

# echo the table for the record
print(f'\n{"RPM":>5} {"m/s":>8} {"kt":>8} {"LPH":>7} {"PCT/HR":>8} {"NM/L":>7}')
for rpm in RPMS:
    kt, lph = kt_at(rpm), lph_at(rpm)
    nml = kt / lph if lph > 0 else 0
    print(f'{rpm:>5} {kt/MS_KT:>8.3f} {kt:>8.3f} {lph:>7.3f} {lph*0.4:>8.3f} {nml:>7.3f}')
